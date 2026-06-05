from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from app import create_app, db
from app.models import (
    DimDate,
    DimMeasure,
    DimPerson,
    FactCheckupMeasurement,
    FactPersonCheckupSnapshot,
)

TARGET_SHEET = "Sheet1"
MAX_HEADER_SCAN_ROWS = 20
DEFAULT_WORKBOOK_PATH = BASE_DIR / "data" / "raw" / "เทคนิค_2565.xlsx"
DATE_CELL = "H2"
LAB_MEASURE_CODES = [
    "Gluc",
    "BUN",
    "CRE",
    "Uric",
    "Chol",
    "TG",
    "AST",
    "ALT",
    "ALK",
    "HDL",
    "LDL",
]
LAB_MEASURE_METADATA = {
    "Gluc": {"measure_name": "Glucose", "category": "Laboratory", "unit": None},
    "BUN": {"measure_name": "Blood Urea Nitrogen", "category": "Laboratory", "unit": None},
    "CRE": {"measure_name": "Creatinine", "category": "Laboratory", "unit": None},
    "Uric": {"measure_name": "Uric Acid", "category": "Laboratory", "unit": None},
    "Chol": {"measure_name": "Cholesterol", "category": "Laboratory", "unit": None},
    "TG": {"measure_name": "Triglycerides", "category": "Laboratory", "unit": None},
    "AST": {"measure_name": "Aspartate Aminotransferase", "category": "Laboratory", "unit": None},
    "ALT": {"measure_name": "Alanine Aminotransferase", "category": "Laboratory", "unit": None},
    "ALK": {"measure_name": "Alkaline Phosphatase", "category": "Laboratory", "unit": None},
    "HDL": {"measure_name": "High-Density Lipoprotein", "category": "Laboratory", "unit": None},
    "LDL": {"measure_name": "Low-Density Lipoprotein", "category": "Laboratory", "unit": None},
}

FIELD_ALIASES = {
    "cms_code": {
        "cmscode",
        "cms",
        "cmsid",
        "รหัสcms",
        "รหัสcms",
        "cms code",
    },
    "employee_id": {
        "employeeid",
        "employee",
        "empid",
        "staffid",
        "รหัสพนักงาน",
        "เลขพนักงาน",
        "รหัส",
    },
    "full_name": {
        "name",
        "fullname",
        "full_name",
        "employee name",
        "ชื่อ",
        "ชื่อสกุล",
        "ชื่อ-สกุล",
        "ชื่อ นามสกุล",
    },
    "age": {"age", "อายุ"},
    "weight": {"weight", "น้ำหนัก"},
    "height": {"height", "ส่วนสูง"},
    "bmi": {"bmi", "bodymassindex"},
    "blood_pressure": {
        "bloodpressure",
        "blood pressure",
        "bp",
        "ความดัน",
        "ความดันโลหิต",
    },
    "Gluc": {"gluc", "glucose"},
    "BUN": {"bun"},
    "CRE": {"cre", "creatinine"},
    "Uric": {"uric", "uricacid"},
    "Chol": {"chol", "cholesterol"},
    "TG": {"tg", "triglyceride", "triglycerides"},
    "AST": {"ast", "sgot"},
    "ALT": {"alt", "sgpt"},
    "ALK": {"alk", "alp", "alkalinephosphatase"},
    "HDL": {"hdl"},
    "LDL": {"ldl"},
}


@dataclass
class ImportStats:
    rows_seen: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    persons_loaded: int = 0
    people_inserted: int = 0
    people_updated: int = 0
    dates_inserted: int = 0
    measures_inserted: int = 0
    snapshots_loaded: int = 0
    snapshots_inserted: int = 0
    snapshots_updated: int = 0
    measurements_loaded: int = 0
    measurements_inserted: int = 0
    measurements_updated: int = 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Load person-level health checkup snapshots from an Excel workbook.",
    )
    parser.add_argument(
        "workbook_path",
        nargs="?",
        default=str(DEFAULT_WORKBOOK_PATH),
        help="Path to the .xlsx workbook to import.",
    )
    return parser.parse_args()


def detect_header_row(preview: pd.DataFrame) -> int:
    best_row_index = 0
    best_score = -1

    for row_index in range(min(len(preview), MAX_HEADER_SCAN_ROWS)):
        row = preview.iloc[row_index]
        non_empty_cells = [cell for cell in row if not is_empty(cell)]
        unique_values = {normalize_cell(cell) for cell in non_empty_cells}
        string_count = sum(isinstance(cell, str) for cell in non_empty_cells)
        score = (len(non_empty_cells) * 3) + (len(unique_values) * 2) + string_count

        if score > best_score:
            best_score = score
            best_row_index = row_index

    return best_row_index


def is_empty(value: Any) -> bool:
    return pd.isna(value) or str(value).strip() == ""


def clean_value(value: Any) -> str:
    if pd.isna(value):
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except TypeError:
            pass
    return str(value).strip()


def normalize_cell(value: Any) -> str:
    return str(value).strip()


def canonicalize_header(header: str) -> str:
    return re.sub(r"[^a-z0-9ก-๙]+", "", header.casefold())


def resolve_columns(column_names: list[str]) -> dict[str, str]:
    available = {canonicalize_header(name): name for name in column_names}
    resolved: dict[str, str] = {}
    missing: list[str] = []

    for field_name, aliases in FIELD_ALIASES.items():
        match = None
        for alias in aliases:
            candidate = available.get(canonicalize_header(alias))
            if candidate:
                match = candidate
                break
        if match:
            resolved[field_name] = match
        else:
            missing.append(field_name)

    required_fields = {
        "cms_code",
        "employee_id",
        "full_name",
        "age",
        "weight",
        "height",
        "bmi",
        "blood_pressure",
        *LAB_MEASURE_CODES,
    }
    required_missing = sorted(required_fields.intersection(missing))
    if required_missing:
        raise ValueError(f"Missing required columns: {', '.join(required_missing)}")

    return resolved


def parse_excel_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    parsed = pd.to_datetime(text, errors="coerce", dayfirst=True)
    if pd.notna(parsed):
        dt = parsed.to_pydatetime().date()
        if dt.year > 2400:
            return dt.replace(year=dt.year - 543)
        return dt

    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if match:
        day, month, year = (int(part) for part in match.groups())
        if year > 2400:
            year -= 543
        elif year < 100:
            year += 2000
        return date(year, month, day)

    raise ValueError(f"Could not parse checkup date from {DATE_CELL}: {value!r}")


def parse_numeric(value: Any) -> float | None:
    if is_empty(value):
        return None

    text = str(value).strip().replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group())


def parse_integer(value: Any) -> int | None:
    numeric = parse_numeric(value)
    if numeric is None:
        return None
    return int(round(numeric))


def parse_blood_pressure(value: Any) -> tuple[float | None, float | None]:
    if is_empty(value):
        return None, None

    parts = re.findall(r"\d+(?:\.\d+)?", str(value))
    if len(parts) < 2:
        return None, None
    return float(parts[0]), float(parts[1])


def split_prefix(full_name: str) -> tuple[str | None, str]:
    known_prefixes = (
        "mr.",
        "mrs.",
        "ms.",
        "miss",
        "dr.",
        "นาย",
        "นาง",
        "นางสาว",
        "ดร.",
    )
    stripped_name = full_name.strip()
    lowered = stripped_name.casefold()
    for prefix in known_prefixes:
        if lowered.startswith(prefix.casefold()):
            remainder = stripped_name[len(prefix):].strip()
            return prefix.strip(), remainder or stripped_name
    return None, stripped_name


def read_checkup_date(workbook_path: Path) -> date:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if TARGET_SHEET not in workbook.sheetnames:
            raise ValueError(f"Worksheet '{TARGET_SHEET}' was not found.")
        sheet = workbook[TARGET_SHEET]
        return parse_excel_date(sheet[DATE_CELL].value)
    finally:
        workbook.close()


def load_sheet(workbook_path: Path) -> tuple[pd.DataFrame, int]:
    preview = pd.read_excel(
        workbook_path,
        sheet_name=TARGET_SHEET,
        header=None,
        nrows=MAX_HEADER_SCAN_ROWS,
    )
    header_row_index = detect_header_row(preview)
    header_row = preview.iloc[header_row_index].tolist()
    column_names = [clean_value(value) or f"unnamed_{index + 1}" for index, value in enumerate(header_row)]

    data_frame = pd.read_excel(
        workbook_path,
        sheet_name=TARGET_SHEET,
        header=header_row_index,
    )
    data_frame.columns = column_names
    data_frame = data_frame.dropna(how="all")
    return data_frame, header_row_index


def get_or_create_person(record: dict[str, Any], stats: ImportStats) -> DimPerson:
    cms_code = str(record["cms_code"]).strip()
    employee_id = str(record["employee_id"]).strip()
    full_name = str(record["full_name"]).strip()
    prefix, clean_name = split_prefix(full_name)

    person = DimPerson.query.filter_by(cms_code=cms_code).one_or_none()
    created = False
    if person is None and employee_id:
        person = DimPerson.query.filter_by(employee_id=employee_id).one_or_none()
    if person is None:
        person = DimPerson(cms_code=cms_code)
        db.session.add(person)
        created = True

    changed = False
    if person.cms_code != cms_code:
        person.cms_code = cms_code
        changed = True
    if employee_id and person.employee_id != employee_id:
        person.employee_id = employee_id
        changed = True
    if person.full_name != clean_name:
        person.full_name = clean_name
        changed = True
    if prefix and person.prefix != prefix:
        person.prefix = prefix
        changed = True

    if created:
        stats.people_inserted += 1
    elif changed:
        stats.people_updated += 1

    return person


def get_or_create_date(checkup_date: date, stats: ImportStats) -> DimDate:
    dim_date = DimDate.query.filter_by(checkup_date=checkup_date).one_or_none()
    if dim_date is None:
        dim_date = DimDate(
            checkup_date=checkup_date,
            checkup_year=checkup_date.year,
        )
        db.session.add(dim_date)
        stats.dates_inserted += 1
    return dim_date


def get_or_create_measure(measure_code: str, stats: ImportStats) -> DimMeasure:
    measure = DimMeasure.query.filter_by(measure_code=measure_code).one_or_none()
    if measure is None:
        metadata = LAB_MEASURE_METADATA[measure_code]
        measure = DimMeasure(
            measure_code=measure_code,
            measure_name=metadata["measure_name"],
            category=metadata["category"],
            unit=metadata["unit"],
        )
        db.session.add(measure)
        stats.measures_inserted += 1
    return measure


def build_snapshot_payload(row: pd.Series, column_map: dict[str, str], source_file: str) -> dict[str, Any]:
    systolic_bp, diastolic_bp = parse_blood_pressure(row[column_map["blood_pressure"]])
    return {
        "age": parse_integer(row[column_map["age"]]),
        "weight": parse_numeric(row[column_map["weight"]]),
        "height": parse_numeric(row[column_map["height"]]),
        "bmi": parse_numeric(row[column_map["bmi"]]),
        "systolic_bp": systolic_bp,
        "diastolic_bp": diastolic_bp,
        "source_file": source_file,
    }


def row_has_identity(cms_code: Any, employee_id: Any, full_name: Any) -> bool:
    return not (is_empty(cms_code) and is_empty(employee_id) and is_empty(full_name))


def upsert_measurements(
    row: pd.Series,
    column_map: dict[str, str],
    person: DimPerson,
    dim_date: DimDate,
    source_file: str,
    measure_dimensions: dict[str, DimMeasure],
    stats: ImportStats,
) -> None:
    for measure_code in LAB_MEASURE_CODES:
        raw_value = row[column_map[measure_code]]
        if is_empty(raw_value):
            continue

        measure = measure_dimensions[measure_code]
        measurement = FactCheckupMeasurement.query.filter_by(
            person_key=person.person_key,
            date_key=dim_date.date_key,
            measure_key=measure.measure_key,
            source_file=source_file,
        ).one_or_none()
        numeric_value = parse_numeric(raw_value)
        raw_text = clean_value(raw_value)

        if measurement is None:
            measurement = FactCheckupMeasurement(
                person_key=person.person_key,
                date_key=dim_date.date_key,
                measure_key=measure.measure_key,
                value_numeric=numeric_value,
                raw_value=raw_text,
                source_file=source_file,
            )
            db.session.add(measurement)
            stats.measurements_inserted += 1
        else:
            measurement.value_numeric = numeric_value
            measurement.raw_value = raw_text
            measurement.source_file = source_file
            stats.measurements_updated += 1

        stats.measurements_loaded += 1


def import_workbook(workbook_path: Path) -> ImportStats:
    stats = ImportStats()
    checkup_date = read_checkup_date(workbook_path)
    data_frame, _ = load_sheet(workbook_path)
    column_map = resolve_columns(data_frame.columns.tolist())
    source_file = workbook_path.name
    dim_date = get_or_create_date(checkup_date, stats)
    measure_dimensions = {
        measure_code: get_or_create_measure(measure_code, stats)
        for measure_code in LAB_MEASURE_CODES
    }
    db.session.flush()

    for _, row in data_frame.iterrows():
        stats.rows_seen += 1

        cms_code = row[column_map["cms_code"]]
        employee_id = row[column_map["employee_id"]]
        full_name = row[column_map["full_name"]]

        if not row_has_identity(cms_code, employee_id, full_name):
            stats.rows_skipped += 1
            continue
        if is_empty(cms_code):
            stats.rows_skipped += 1
            continue

        person = get_or_create_person(
            {
                "cms_code": clean_value(cms_code),
                "employee_id": clean_value(employee_id),
                "full_name": clean_value(full_name),
            },
            stats,
        )
        db.session.flush()
        stats.persons_loaded += 1

        snapshot_payload = build_snapshot_payload(row, column_map, source_file)
        snapshot = FactPersonCheckupSnapshot.query.filter_by(
            person_key=person.person_key,
            date_key=dim_date.date_key,
            source_file=source_file,
        ).one_or_none()

        if snapshot is None:
            snapshot = FactPersonCheckupSnapshot(
                person_key=person.person_key,
                date_key=dim_date.date_key,
                **snapshot_payload,
            )
            db.session.add(snapshot)
            stats.snapshots_inserted += 1
        else:
            for key, value in snapshot_payload.items():
                setattr(snapshot, key, value)
            stats.snapshots_updated += 1
        stats.snapshots_loaded += 1

        upsert_measurements(
            row=row,
            column_map=column_map,
            person=person,
            dim_date=dim_date,
            source_file=source_file,
            measure_dimensions=measure_dimensions,
            stats=stats,
        )

        stats.rows_imported += 1

    db.session.commit()
    return stats


def print_report(workbook_path: Path, checkup_date: date, stats: ImportStats) -> None:
    print("Checkup Import Report")
    print("====================")
    print(f"Workbook: {workbook_path.name}")
    print(f"Path: {workbook_path}")
    print(f"Sheet: {TARGET_SHEET}")
    print(f"Checkup date ({DATE_CELL}): {checkup_date.isoformat()}")
    print("")
    print("Statistics")
    print("----------")
    print(f"Persons loaded: {stats.persons_loaded}")
    print(f"Snapshots loaded: {stats.snapshots_loaded}")
    print(f"Measurements loaded: {stats.measurements_loaded}")
    print(f"Rows seen: {stats.rows_seen}")
    print(f"Rows imported: {stats.rows_imported}")
    print(f"Rows skipped: {stats.rows_skipped}")
    print(f"dim_person inserted: {stats.people_inserted}")
    print(f"dim_person updated: {stats.people_updated}")
    print(f"dim_date inserted: {stats.dates_inserted}")
    print(f"dim_measure inserted: {stats.measures_inserted}")
    print(f"fact_person_checkup_snapshot inserted: {stats.snapshots_inserted}")
    print(f"fact_person_checkup_snapshot updated: {stats.snapshots_updated}")
    print(f"fact_checkup_measurement inserted: {stats.measurements_inserted}")
    print(f"fact_checkup_measurement updated: {stats.measurements_updated}")


def main() -> int:
    args = parse_args()
    workbook_path = Path(args.workbook_path).expanduser().resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 1

    app = create_app()
    with app.app_context():
        db.create_all()
        checkup_date = read_checkup_date(workbook_path)
        stats = import_workbook(workbook_path)
        print_report(workbook_path, checkup_date, stats)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
